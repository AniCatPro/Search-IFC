import os
import sqlite3
from datetime import datetime
import getpass
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from rapidfuzz import fuzz
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import letter, landscape
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from tkinter import Tk, filedialog, Label, Button, Entry, messagebox, IntVar, Checkbutton, ttk, TclError

conn = None
cursor = None


def connect_to_database(db_path):
    global conn, cursor
    if conn:
        conn.close()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY,
            parent_folder TEXT,
            path TEXT,
            filename TEXT,
            last_modified TEXT,
            created_by TEXT
        )
    ''')
    conn.commit()
    messagebox.showinfo("Информация", "Подключение к базе данных успешно!")


def update_file_in_db(parent_folder, path, filename, last_modified, created_by):
    cursor.execute('''
        INSERT OR REPLACE INTO files (parent_folder, path, filename, last_modified, created_by)
        VALUES (?, ?, ?, ?, ?)
    ''', (parent_folder, path, filename, last_modified, created_by))
    conn.commit()
    update_table()


def are_filenames_similar(name1, name2, threshold=80):
    base1, _ = os.path.splitext(name1)
    base2, _ = os.path.splitext(name2)
    similarity = fuzz.ratio(base1, base2)
    return similarity >= threshold


def apply_highlighting():
    for row in tree.get_children():
        filename, parent_folder, path, last_modified, _ = tree.item(row, "values")
        tag = 'rvt' if filename.endswith('.rvt') else 'ifc' if filename.endswith('.ifc') else 'dwg'

        # Applying basic highlighting
        tree.item(row, tags=(tag,))

    try:
        threshold = similarity_threshold.get()
    except TclError:
        return

    rows = cursor.execute("SELECT * FROM files").fetchall()
    file_groups = {}
    for row in rows:
        parent_folder, path, filename, last_modified, created_by = row[1:]
        key = (parent_folder, last_modified[:10])
        if key not in file_groups:
            file_groups[key] = []
        file_groups[key].append(row)

    # Checking for similarity and applying highlight if needed
    for group in file_groups.values():
        for i, file1 in enumerate(group):
            for j, file2 in enumerate(group):
                if i < j:
                    if are_filenames_similar(file1[3], file2[3], threshold) and file1[4][:10] == file2[4][:10]:
                        if (file1[3].endswith(".rvt") and file2[3].endswith(".ifc")) or \
                                (file1[3].endswith(".ifc") and file2[3].endswith(".rvt")) or \
                                (file1[3].endswith(".dwg") and file2[3].endswith(".ifc")) or \
                                (file1[3].endswith(".ifc") and file2[3].endswith(".dwg")):
                            for f in (file1, file2):
                                for row in tree.get_children():
                                    if tree.item(row, "values")[2] == f[2]:
                                        current_tags = tree.item(row, "tags")
                                        tree.item(row, tags=current_tags + ("highlight",))


def update_table():
    for row in tree.get_children():
        tree.delete(row)
    rows = cursor.execute("SELECT filename, parent_folder, path, last_modified, created_by FROM files").fetchall()
    rows = sorted(rows, key=lambda x: (x[1], x[3], x[0]))
    for row in rows:
        filename, parent_folder, path, last_modified, created_by = row
        tag = 'rvt' if filename.endswith('.rvt') else ('dwg' if filename.endswith('.dwg') else 'ifc')
        tree.insert('', 'end', values=row, tags=(tag,))
    apply_highlighting()


def wrap_text(text, max_width, font, font_size, pdf_canvas):
    words = text.split(' ')
    lines = []
    current_line = words[0]
    for word in words[1:]:
        if pdf_canvas.stringWidth(current_line + ' ' + word, font, font_size) < max_width:
            current_line += ' ' + word
        else:
            lines.append(current_line)
            current_line = word
    lines.append(current_line)
    return lines


def create_pdf_report():
    highlighted_files = []
    for row in tree.get_children():
        item = tree.item(row)
        if "highlight" in item["tags"]:
            highlighted_files.append(item["values"])
    if not highlighted_files:
        messagebox.showwarning("Предупреждение", "Нет совпадающих файлов для генерации отчета.")
        return
    save_folder = filedialog.askdirectory()
    if not save_folder:
        return
    root_folder_name = os.path.basename(folder_path.get())
    report_name = f"{root_folder_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf"
    pdf_path = os.path.join(save_folder, report_name)
    font_path = r"C:\Windows\Fonts\arial.ttf"
    pdfmetrics.registerFont(TTFont('Arial', font_path))
    c = canvas.Canvas(pdf_path, pagesize=landscape(letter))
    width, height = landscape(letter)
    c.setFont("Arial", 14)
    c.drawString(30, height - 30, f"Отчет по файлам в '{root_folder_name}'")
    c.setFont("Arial", 10)
    c.drawString(30, height - 50, f"Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y_position = height - 70
    files_by_parent_folder = {}
    for file in highlighted_files:
        filename, parent_folder, path, last_modified, created_by = file
        if parent_folder not in files_by_parent_folder:
            files_by_parent_folder[parent_folder] = []
        files_by_parent_folder[parent_folder].append(file)
    for parent_folder, files in files_by_parent_folder.items():
        c.setFont("Arial", 12)
        c.drawString(30, y_position, f"Папка: {parent_folder}")
        y_position -= 20
        c.setFont("Arial", 10)
        c.drawString(30, y_position, "Имя файла")
        c.drawString(200, y_position, "Последнее изм.")
        c.drawString(300, y_position, "Путь")
        y_position -= 20
        for file in files:
            filename, parent_folder, path, last_modified, created_by = file
            c.drawString(30, y_position, filename)
            c.drawString(200, y_position, last_modified)
            path_max_width = width - 440
            path_lines = wrap_text(path, path_max_width, "Arial", 10, c)
            for line in path_lines:
                c.drawString(300, y_position, line)
                y_position -= 12
            if len(path_lines) == 1 and c.stringWidth(path_lines[0], "Arial", 10) <= path_max_width:
                c.setFillColor(colors.blue)
                c.linkURL(f"file://{path}", (300, y_position + 12, width - 30, y_position + 24), relative=0)
                c.setFillColor(colors.black)
            y_position -= 8
            if y_position < 40:
                c.showPage()
                c.setFont("Arial", 10)
                y_position = height - 30
                c.drawString(30, y_position, "Имя файла")
                c.drawString(200, y_position, "Последнее изм.")
                c.drawString(300, y_position, "Путь")
                y_position -= 20
    c.save()
    messagebox.showinfo("Информация", f"Отчет сохранен в {pdf_path}")


def export_to_excel():
    highlighted_files = []
    for row in tree.get_children():
        item = tree.item(row)
        if "highlight" in item["tags"]:
            highlighted_files.append(item["values"])
    if not highlighted_files:
        messagebox.showwarning("Предупреждение", "Нет выделенных файлов для экспорта.")
        return

    save_folder = filedialog.askdirectory()
    if not save_folder:
        return

    root_folder_name = os.path.basename(folder_path.get())
    report_name = f"{root_folder_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    save_path = os.path.join(save_folder, report_name)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Отчет по файлам"

    headers = ["Имя файла", "Родительская Папка", "Путь", "Последнее Изменение", "Создано"]
    sheet.append(headers)

    for header_cell in sheet[1]:
        header_cell.font = Font(bold=True)

    for file in highlighted_files:
        sheet.append(file)

    last_row = sheet.max_row
    last_column = sheet.max_column
    table_range = f"A1:{chr(64 + last_column)}{last_row}"

    table = Table(displayName="FilesTable", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style

    sheet.add_table(table)

    adjust_column_width(sheet, headers)

    green_fill = PatternFill(start_color="C4DFB1", end_color="C4DFB1", fill_type="solid")

    sheet.conditional_formatting.add(f"A2:{chr(64 + last_column)}{last_row}",
                                     CellIsRule(operator='equal', formula=[f'"highlight"'], fill=green_fill))

    workbook.save(save_path)

    messagebox.showinfo("Информация", f"Excel отчет сохранен в {save_path}")

def adjust_column_width(sheet, headers):
    for col_idx, col_header in enumerate(headers, start=1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        max_length = len(col_header)

        for cell in sheet[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 1

def copy_path():
    selected_item = tree.focus()
    if not selected_item:
        messagebox.showwarning("Предупреждение", "Файл не выбран!")
        return
    path = tree.item(selected_item, "values")[2]
    root.clipboard_clear()
    root.clipboard_append(path)
    messagebox.showinfo("Информация", f"Путь скопирован в буфер обмена:\n{path}")


def scan_work_folders(root_folder):
    file_extensions = []
    if monitor_rvt_ifc.get():
        file_extensions.extend(['.rvt', '.ifc'])
    if monitor_dwg_ifc.get():
        file_extensions.extend(['.dwg', '.ifc'])
    for dirpath, dirnames, filenames in os.walk(root_folder):
        if os.path.basename(dirpath) == "Работа":
            parent_folder = os.path.basename(os.path.dirname(dirpath))
            for filename in filenames:
                if any(filename.endswith(ext) for ext in file_extensions):
                    file_path = os.path.join(dirpath, filename)
                    last_modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                    created_by = getpass.getuser()
                    update_file_in_db(parent_folder, file_path, filename, last_modified, created_by)


class FileMonitorHandler(FileSystemEventHandler):
    def process_file(self, event):
        file_extensions = []
        if monitor_rvt_ifc.get():
            file_extensions.extend(['.rvt', '.ifc'])
        if monitor_dwg_ifc.get():
            file_extensions.extend(['.dwg', '.ifc'])
        if not event.is_directory:
            filename = os.path.basename(event.src_path)
            parent_folder = os.path.basename(os.path.dirname(os.path.dirname(event.src_path)))
            if any(filename.endswith(ext) for ext in file_extensions) and "Работа" in os.path.dirname(event.src_path):
                last_modified = datetime.fromtimestamp(os.path.getmtime(event.src_path)).strftime('%Y-%m-%d %H:%M:%S')
                created_by = getpass.getuser()
                update_file_in_db(parent_folder, event.src_path, filename, last_modified, created_by)

    def on_modified(self, event):
        self.process_file(event)

    def on_created(self, event):
        self.process_file(event)

    def on_deleted(self, event):
        if not event.is_directory:
            cursor.execute("DELETE FROM files WHERE path = ?", (event.src_path,))
            conn.commit()
            update_table()


def select_folder():
    selected_folder = filedialog.askdirectory()
    if selected_folder:
        folder_path.delete(0, 'end')
        folder_path.insert(0, selected_folder)


def select_db_folder():
    selected_db_folder = filedialog.askdirectory()
    if selected_db_folder:
        db_folder_path.delete(0, 'end')
        db_folder_path.insert(0, selected_db_folder)


def select_existing_db():
    existing_db = filedialog.askopenfilename(filetypes=[("SQLite Database", "*.db")])
    if existing_db:
        old_db_path.delete(0, 'end')
        old_db_path.insert(0, existing_db)


def start_monitoring():
    if use_existing_db.get():
        existing_db = old_db_path.get()
        if not existing_db:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите старую базу данных для использования.")
            return
        connect_to_database(existing_db)
        update_table()
    else:
        path = folder_path.get()
        db_folder = db_folder_path.get()
        if not path:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите папку для мониторинга.")
            return
        if not db_folder:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите папку для сохранения базы данных.")
            return

        folder_name = os.path.basename(path)
        db_name = f"BD_{folder_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.db"
        db_path = os.path.join(db_folder, db_name)

        # Создание новой базы данных
        connect_to_database(db_path)

        # Запуск сканирования и наблюдения
        scan_work_folders(path)
        event_handler = FileMonitorHandler()
        observer = Observer()
        observer.schedule(event_handler, path=path, recursive=True)
        observer.start()


def toggle_db_mode():
    if use_existing_db.get():
        folder_button.config(state='disabled')
        db_folder_button.config(state='disabled')
        existing_db_button.config(state='normal')
        old_db_path.config(state='normal')
        for row in tree.get_children():
            tree.delete(row)
    else:
        folder_button.config(state='normal')
        db_folder_button.config(state='normal')
        existing_db_button.config(state='disabled')
        old_db_path.config(state='disabled')


def toggle_compare_mode():
    if compare_with_old.get():
        existing_db_button.config(state='normal')
        compare_button.config(state='normal')
    else:
        existing_db_button.config(state='disabled')
        compare_button.config(state='disabled')


def compare_databases():
    path = folder_path.get()
    db_folder = db_folder_path.get()
    if not path:
        messagebox.showerror("Ошибка", "Пожалуйста, выберите папку для мониторинга.")
        return
    if not db_folder:
        messagebox.showerror("Ошибка", "Пожалуйста, выберите папку для сохранения базы данных.")
        return

    folder_name = os.path.basename(path)
    db_name = f"BD_{folder_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.db"
    db_path = os.path.join(db_folder, db_name)

    # Создание и подключение к новой базе данных
    connect_to_database(db_path)

    old_db_path_value = old_db_path.get()
    if not old_db_path_value:
        messagebox.showerror("Ошибка", "Пожалуйста, выберите старую базу данных для сравнения.")
        return

    try:
        # Подключаемся к старой базе данных
        old_conn = sqlite3.connect(old_db_path_value)
        old_cursor = old_conn.cursor()
        old_files = old_cursor.execute("SELECT filename, parent_folder, path, last_modified FROM files").fetchall()
        old_conn.close()

        # Получаем данные из текущей (новой) базы данных
        new_files = cursor.execute("SELECT filename, parent_folder, path, last_modified FROM files").fetchall()

        # Создаем словари для удобного сравнения
        old_files_dict = {file[2]: file for file in old_files}
        new_files_dict = {file[2]: file for file in new_files}

        # Очищаем древовидный виджет
        for row in tree.get_children():
            tree.delete(row)

        # Итерируем по новым файлам и определяем их статус
        for new_path, new_file in new_files_dict.items():
            if new_path in old_files_dict:
                old_file = old_files_dict[new_path]
                # Обновлено: Проверяем, изменилось ли время последнего изменения
                if old_file[3] != new_file[3]:
                    tag = "changed"
                else:
                    tag = "normal"
            else:
                tag = "new"

            tree.insert('', 'end', values=new_file, tags=(tag,))

        # Итерируем по старым файлам, которых нет в новых
        for old_path, old_file in old_files_dict.items():
            if old_path not in new_files_dict:
                tree.insert('', 'end', values=old_file, tags=("missing",))

        # Настройка тегов для отображения изменений
        tree.tag_configure("changed", background="orange")  # Измененные файлы
        tree.tag_configure("new", background="lightgreen")  # Новые файлы
        tree.tag_configure("missing", background="red")  # Удаленные файлы

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось выполнить сравнение: {e}")


def update_highlighting():
    apply_highlighting()


root = Tk()
root.title("Ищем актуальные IFC")

use_existing_db = IntVar(value=0)
use_existing_checkbox = Checkbutton(root, text="Использовать старую БД", variable=use_existing_db,
                                    command=toggle_db_mode)
use_existing_checkbox.grid(row=0, column=0, padx=10, pady=10, sticky="w")

compare_with_old = IntVar(value=0)
compare_with_checkbox = Checkbutton(root, text="Сравнить со старой БД", variable=compare_with_old,
                                    command=toggle_compare_mode)
compare_with_checkbox.grid(row=0, column=1, padx=10, pady=10, sticky="w")

folder_label = Label(root, text="Выбрать папку:")
folder_label.grid(row=1, column=0, padx=10, pady=10)

folder_path = Entry(root, width=50)
folder_path.grid(row=1, column=1, padx=10, pady=10)

folder_button = Button(root, text="Выбрать", command=select_folder)
folder_button.grid(row=1, column=2, padx=10, pady=10)

db_folder_label = Label(root, text="Выбрать папку для базы данных:")
db_folder_label.grid(row=2, column=0, padx=10, pady=10)

db_folder_path = Entry(root, width=50)
db_folder_path.grid(row=2, column=1, padx=10, pady=10)

db_folder_button = Button(root, text="Выбрать", command=select_db_folder)
db_folder_button.grid(row=2, column=2, padx=10, pady=10)

existing_db_label = Label(root, text="Выбрать файл старой базы данных:")
existing_db_label.grid(row=3, column=0, padx=10, pady=10)

old_db_path = Entry(root, width=50, state='disabled')
old_db_path.grid(row=3, column=1, padx=10, pady=10)

existing_db_button = Button(root, text="Выбрать", command=select_existing_db, state='disabled')
existing_db_button.grid(row=3, column=2, padx=10, pady=10)

compare_button = Button(root, text="Сравнить", command=compare_databases, state='disabled')
compare_button.grid(row=4, column=1, padx=10, pady=10)

similarity_threshold_label = Label(root, text="% схожести:")
similarity_threshold_label.grid(row=5, column=0, padx=10, pady=10)

similarity_threshold = IntVar(value=80)
similarity_threshold_entry = Entry(root, textvariable=similarity_threshold, width=5)
similarity_threshold_entry.grid(row=5, column=1, padx=10, pady=10, sticky="w")

apply_button = Button(root, text="Применить %", command=update_highlighting)
apply_button.grid(row=5, column=2, padx=10, pady=10)

monitor_rvt_ifc = IntVar(value=1)
monitor_dwg_ifc = IntVar(value=0)

rvt_ifc_checkbox = Checkbutton(root, text=".rvt и .ifc", variable=monitor_rvt_ifc)
rvt_ifc_checkbox.grid(row=6, column=0, padx=10, pady=10)

dwg_ifc_checkbox = Checkbutton(root, text=".dwg и .ifc", variable=monitor_dwg_ifc)
dwg_ifc_checkbox.grid(row=6, column=1, padx=10, pady=10)

start_button = Button(root, text="Начать мониторинг", command=start_monitoring)
start_button.grid(row=7, column=0, columnspan=3, padx=10, pady=20)

columns = ("Имя файла", "Родительская Папка", "Путь", "Последнее Изменение", "Создано")
tree = ttk.Treeview(root, columns=columns, show="headings")
tree.heading("Имя файла", text="Имя файла")
tree.heading("Родительская Папка", text="Папка")
tree.heading("Путь", text="Путь")
tree.heading("Последнее Изменение", text="Последнее изм.")
tree.heading("Создано", text="Кем создано")
tree.grid(row=8, column=0, columnspan=3, padx=10, pady=20, sticky="nsew")

root.grid_rowconfigure(8, weight=1)
root.grid_columnconfigure(1, weight=1)

tree.tag_configure("highlight", background="lightgreen")
tree.tag_configure("rvt", background="lightblue")
tree.tag_configure("dwg", background="lavender")
tree.tag_configure("ifc", background="lightyellow")
tree.tag_configure("changed", background="orange")
tree.tag_configure("new", background="lightgreen")
tree.tag_configure("missing", background="red")

pdf_button = Button(root, text="Отчет PDF", command=create_pdf_report)
pdf_button.grid(row=9, column=0, padx=10, pady=10)

excel_button = Button(root, text="Отчет Excel", command=export_to_excel)
excel_button.grid(row=9, column=1, padx=10, pady=10)

copy_button = Button(root, text="Скопировать путь", command=copy_path)
copy_button.grid(row=9, column=2, padx=10, pady=10)

root.mainloop()

if conn:
    conn.close()